#withfirefox_fetch
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Geckodriver yolunu belirtin
geckodriver_path = '/Users/sukru/myenv/bin/geckodriver'
service = Service(executable_path=geckodriver_path)

# Firefox profil yolunu belirtin
profile_path = '/Users/sukru/Library/Application Support/Firefox/Profiles/s1p7nlrw.default-release-1730826088264'
profile = webdriver.FirefoxProfile(profile_path)

# Firefox seçeneklerini yapılandır (profili ekle)
options = Options()
options.profile = profile

# Tarayıcıyı başlat (profil ile)
driver = webdriver.Firefox(service=service, options=options)

# Tarayıcıyı başlat ve login için bekle
driver.get("https://google.com")
print("Login bilgilerinizi girmeniz için 15 saniye bekleniyor...")
time.sleep(2)  # Login işlemleri için bekleme

# Excel'den URL'leri yükle
input_excel_path = "/Users/sukru/Downloads/pandora_all_url.xlsx"
data = pd.read_excel(input_excel_path)

# Yeni CSS seçicilerip
selectors = {
    "anchor": "h1.product-name",
    "price1": ".price-attribute .prices .price",
    "stock": "[data-link=addToCart]"
}

# Çıktı verisini saklamak için bir liste
output_data = []
output_excel_path = "/Users/sukru/Downloads/firefox_pandora_results.xlsx"

# URL'leri sırasıyla ziyaret et ve bilgiyi topla
for index, row in data.iterrows():
    url = row['URL']
    driver.get(url)
    time.sleep(2)  # Sayfanın yüklenmesi için kısa bekleme

    row_data = {'URL': url}  # Her URL için toplanan veriler
    all_failed = True  # Tüm seçiciler başarısız olursa kontrolü

    for key, selector in selectors.items():
        try:
            element = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, selector))
            )
            content = element.text.strip() if element.text.strip() else "Veri bulunamadı"
            all_failed = False  # En az bir seçici çalıştıysa başarısız sayma
        except:
            content = "Hata: Eleman bulunamadı"

        row_data[key] = content

    # Eğer tüm seçiciler başarısız olduysa kaydetmeden sonraki URL'ye geç
    if all_failed:
        print(f"URL {url} için hiçbir veri bulunamadı, atlanıyor.")
        continue

    output_data.append(row_data)
    temp_df = pd.DataFrame(output_data)
    temp_df.to_excel(output_excel_path, index=False)

    # Biçimlendirme için openpyxl ile işle
    workbook = load_workbook(output_excel_path)
    sheet = workbook.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    if index == 0:
        for col_num, key in enumerate(selectors.keys(), start=2):
            sheet.cell(row=1, column=col_num, value=key)

    for row_num in range(2, sheet.max_row + 1):
        for col_num in range(2, len(selectors) + 2):
            cell = sheet.cell(row=row_num, column=col_num)
            if cell.value in ["Veri bulunamadı", "Hata: Eleman bulunamadı"]:
                cell.fill = red_fill

    workbook.save(output_excel_path)
    print(f"Anlık veriler {output_excel_path} dosyasına kaydedildi.")

print(f"İşlem tamamlandı! Çıktılar {output_excel_path} dosyasına kaydedildi.")

driver.quit()